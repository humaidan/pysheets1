import openpyxl
import warnings
from config import CONFIG


# Filter the specific warnings
warnings.filterwarnings("ignore", category=UserWarning, message="Unknown extension is not supported and will be removed")
warnings.filterwarnings("ignore", category=UserWarning, message="Cannot parse header or footer so it will be ignored")

debug = False


####################### main #######################
if __name__ == "__main__":
    print("Opening source sheet ...")

    try:
        wb = openpyxl.load_workbook(CONFIG["excelfile"])

    except Exception as e:
        print(f"Source workbook [{CONFIG['excelfile']}] error: {e}")
        raise  # Raise an exception to halt further code execution

    print("Parsing and writing data ...")

    sheets = 0
    start_row = CONFIG["start_row"]
    sheet1 = wb[CONFIG["sheetname_source"]]
    sheet2 = wb[CONFIG["sheetname_dest"]]

    for row_no, row_val in enumerate(sheet1.iter_rows(min_row=start_row, values_only=True)):
        if row_val[CONFIG["sysname_col"]] is None:
            print("Reached empty service:")
            print(row_val)
            break

        if debug:
            print(f"- looping through record #{row_no+start_row}:")
        newsheet = wb.copy_worksheet(sheet2)
        newsheet.sheet_view.showGridLines = False

        newsheet.conditional_formatting._cf_rules = sheet2.conditional_formatting._cf_rules.copy()

        if debug:
            print(f"\t - set title to: {row_val[CONFIG['sysname_col']]} + {CONFIG['sysname_col_append']}:")

        newtitle = row_val[CONFIG["sysname_col"]].rstrip()
        if newtitle in CONFIG["short_titles"]:
            newtitle = CONFIG["short_titles"][newtitle]

        newtitle = newtitle.replace("- ", "")
        max_title_chars = CONFIG["max_title_chars"] - len(CONFIG["sysname_col_append"])
        if len(newtitle) > max_title_chars:
            newtitle = newtitle[: max_title_chars - 1] + "."

        newsheet.title = newtitle + CONFIG["sysname_col_append"]

        for target_cell, source_col_str in CONFIG["field_mapping"].items():
            # Read the value in "All Services" sheet in current row + config column
            source_value = ""
            if isinstance(source_col_str, str):
                source_col_val = openpyxl.utils.column_index_from_string(source_col_str)
                if debug:
                    print(f"\t - sheet1.cell(row={row_no+start_row}, column={source_col_val}).value")
                source_value = sheet1.cell(row=row_no + start_row, column=source_col_val).value
                if source_value is None:
                    source_value = ""

                if target_cell in CONFIG["pre_fills"] and source_value:
                    source_value = CONFIG["pre_fills"][target_cell] + str(source_value)

                if debug:
                    print(f"\t - reading AllSystems[row={row_no + start_row}, column={source_col_val}]  =  {source_value}.")

            else:  # list of str
                if debug:
                    print(f"\t - detected a source list {source_col_str}:")

                source_value = ""
                for col in source_col_str:
                    source_col_val = openpyxl.utils.column_index_from_string(col)
                    cell = sheet1.cell(row=row_no + start_row, column=source_col_val).value
                    if cell is not None:
                        source_value += "- " + str(cell) + "\n"

            # Write value in new sheet
            if debug:
                print(f"\t=Setting {newsheet.title} sheet cell [{target_cell}] to value: {source_value}.")
            if debug:
                print()
            newsheet[target_cell] = str(source_value).rstrip()

        newsheet[target_cell].alignment = newsheet[target_cell].alignment.copy(wrap_text=True)
        # newsheet[target_cell].alignment = openpyxl.styles.Alignment(horizontal="center")
        sheets += 1

    print("Cleaning ...")
    wb.save(CONFIG["modified_excelfile"])

    print(f"{sheets} new sheets added.")
    print()
